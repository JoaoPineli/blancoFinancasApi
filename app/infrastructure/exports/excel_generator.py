"""Excel report generator using OpenPyXL."""

from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class ExcelReportGenerator:
    """Generator for Excel reports.

    Generates reports for:
    - Fluxo de Caixa (Cash Flow)
    - Conciliação (Reconciliation)
    """

    def __init__(self) -> None:
        """Initialize Excel report generator."""
        # Define styles
        self._header_font = Font(bold=True, color="FFFFFF")
        self._header_fill = PatternFill(
            start_color="1a365d", end_color="1a365d", fill_type="solid"
        )
        self._border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self._money_format = 'R$ #,##0.00'
        self._date_format = 'DD/MM/YYYY'
        self._datetime_format = 'DD/MM/YYYY HH:MM'

    def generate_cash_flow_report(
        self,
        transactions: List[Dict[str, Any]],
        start_date: datetime,
        end_date: datetime,
    ) -> BytesIO:
        """Generate cash flow report.

        Args:
            transactions: List of transaction data
            start_date: Report start date
            end_date: Report end date

        Returns:
            BytesIO buffer with Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Fluxo de Caixa"

        # Title
        ws.merge_cells("A1:G1")
        ws["A1"] = f"Relatório de Fluxo de Caixa - {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Headers
        headers = [
            "Data",
            "Tipo",
            "Cliente",
            "Descrição",
            "Entrada",
            "Saída",
            "Status",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.border = self._border
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        row = 4
        total_inflow = Decimal("0")
        total_outflow = Decimal("0")

        for txn in transactions:
            ws.cell(row=row, column=1, value=txn["date"]).number_format = self._datetime_format
            ws.cell(row=row, column=2, value=txn["type"])
            ws.cell(row=row, column=3, value=txn["client_name"])
            ws.cell(row=row, column=4, value=txn["description"])

            # Inflow (deposits)
            if txn["type"] in ["deposit", "yield"]:
                amount = Decimal(txn["amount"]) / Decimal(100)
                ws.cell(row=row, column=5, value=float(amount)).number_format = self._money_format
                total_inflow += amount
            else:
                ws.cell(row=row, column=5, value="")

            # Outflow (withdrawals)
            if txn["type"] == "withdrawal":
                amount = Decimal(txn["amount"]) / Decimal(100)
                ws.cell(row=row, column=6, value=float(amount)).number_format = self._money_format
                total_outflow += amount
            else:
                ws.cell(row=row, column=6, value="")

            ws.cell(row=row, column=7, value=txn["status"])

            # Apply borders
            for col in range(1, 8):
                ws.cell(row=row, column=col).border = self._border

            row += 1

        # Totals row
        ws.cell(row=row + 1, column=4, value="TOTAIS").font = Font(bold=True)
        ws.cell(row=row + 1, column=5, value=float(total_inflow)).number_format = self._money_format
        ws.cell(row=row + 1, column=5).font = Font(bold=True)
        ws.cell(row=row + 1, column=6, value=float(total_outflow)).number_format = self._money_format
        ws.cell(row=row + 1, column=6).font = Font(bold=True)

        # Auto-adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_clients_report(
        self,
        clients: List[Dict[str, Any]],
        start_date: datetime,
        end_date: datetime,
    ) -> BytesIO:
        """Generate clients report.

        Args:
            clients: List of client dicts with integer centavos fields.
            start_date: Report start date.
            end_date: Report end date.

        Returns:
            BytesIO buffer with Excel file.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Clientes"

        ws.merge_cells("A1:J1")
        ws["A1"] = (
            f"Relatório de Clientes"
            f" - {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"
        )
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = [
            "Nome", "Email", "CPF", "Telefone", "Status",
            "Saldo (centavos)", "Total Investido (centavos)",
            "Rendimento Total (centavos)", "Fundo Garantidor (centavos)",
            "Criado em",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.border = self._border
            cell.alignment = Alignment(horizontal="center")

        for row_idx, c in enumerate(clients, start=4):
            ws.cell(row=row_idx, column=1, value=c["name"])
            ws.cell(row=row_idx, column=2, value=c["email"])
            ws.cell(row=row_idx, column=3, value=c.get("cpf", ""))
            ws.cell(row=row_idx, column=4, value=c.get("phone", ""))
            ws.cell(row=row_idx, column=5, value=c["status"])
            # Store centavos as integers — no float
            ws.cell(row=row_idx, column=6, value=c.get("balance_cents", 0))
            ws.cell(row=row_idx, column=7, value=c.get("total_invested_cents", 0))
            ws.cell(row=row_idx, column=8, value=c.get("total_yield_cents", 0))
            ws.cell(row=row_idx, column=9, value=c.get("fundo_garantidor_cents", 0))
            ws.cell(row=row_idx, column=10, value=c["created_at"])
            for col in range(1, 11):
                ws.cell(row=row_idx, column=col).border = self._border

        column_widths = [30, 35, 14, 14, 12, 20, 25, 25, 25, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_transactions_report(
        self,
        transactions: List[Dict[str, Any]],
        start_date: datetime,
        end_date: datetime,
    ) -> BytesIO:
        """Generate transactions report.

        Args:
            transactions: List of transaction dicts.
            start_date: Report start date.
            end_date: Report end date.

        Returns:
            BytesIO buffer with Excel file.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Transações"

        ws.merge_cells("A1:I1")
        ws["A1"] = (
            f"Relatório de Transações"
            f" - {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"
        )
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = [
            "ID", "Cliente", "Data", "Confirmado em",
            "Tipo", "Status", "Valor (centavos)", "ID Pix", "Descrição",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.border = self._border
            cell.alignment = Alignment(horizontal="center")

        for row_idx, t in enumerate(transactions, start=4):
            ws.cell(row=row_idx, column=1, value=str(t["id"]))
            ws.cell(row=row_idx, column=2, value=t["client_name"])
            ws.cell(row=row_idx, column=3, value=t["created_at"]).number_format = self._datetime_format
            confirmed = t.get("confirmed_at")
            ws.cell(row=row_idx, column=4, value=confirmed if confirmed else "").number_format = self._datetime_format
            ws.cell(row=row_idx, column=5, value=t["transaction_type"])
            ws.cell(row=row_idx, column=6, value=t["status"])
            ws.cell(row=row_idx, column=7, value=t["amount_cents"])  # integer centavos
            ws.cell(row=row_idx, column=8, value=t.get("pix_transaction_id", ""))
            ws.cell(row=row_idx, column=9, value=t.get("description", ""))
            for col in range(1, 10):
                ws.cell(row=row_idx, column=col).border = self._border

        column_widths = [36, 30, 18, 18, 16, 12, 20, 28, 40]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_yields_report(
        self,
        yields: List[Dict[str, Any]],
        start_date: datetime,
        end_date: datetime,
    ) -> BytesIO:
        """Generate yields report based on AuditLog YIELD_CREDITED entries.

        Args:
            yields: List of yield dicts derived from AuditLog details.
            start_date: Report start date.
            end_date: Report end date.

        Returns:
            BytesIO buffer with Excel file.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Rendimentos"

        ws.merge_cells("A1:K1")
        ws["A1"] = (
            f"Relatório de Rendimentos"
            f" - {start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"
        )
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        headers = [
            "Usuário", "Série SGS", "Período De", "Período Até",
            "Taxa Efetiva", "Principal (centavos)", "Rendimento (centavos)",
            "Nº Parcela", "ID Assinatura", "ID Depósito Principal", "Data",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.border = self._border
            cell.alignment = Alignment(horizontal="center")

        for row_idx, y in enumerate(yields, start=4):
            ws.cell(row=row_idx, column=1, value=y["user_name"])
            ws.cell(row=row_idx, column=2, value=y["sgs_series_id"])
            ws.cell(row=row_idx, column=3, value=y["yield_period_from"])
            ws.cell(row=row_idx, column=4, value=y["yield_period_to"])
            ws.cell(row=row_idx, column=5, value=str(y["effective_rate"]))
            ws.cell(row=row_idx, column=6, value=y["principal_cents"])  # integer centavos
            ws.cell(row=row_idx, column=7, value=y["yield_cents"])      # integer centavos
            ws.cell(row=row_idx, column=8, value=y["installment_number"])
            ws.cell(row=row_idx, column=9, value=str(y["subscription_id"]))
            ws.cell(row=row_idx, column=10, value=str(y["principal_deposit_id"]))
            ws.cell(row=row_idx, column=11, value=y["created_at"]).number_format = self._datetime_format
            for col in range(1, 12):
                ws.cell(row=row_idx, column=col).border = self._border

        column_widths = [30, 12, 12, 12, 20, 22, 22, 12, 36, 36, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def generate_reconciliation_report(
        self,
        transactions: List[Dict[str, Any]],
    ) -> BytesIO:
        """Generate reconciliation report.

        Args:
            transactions: List of transaction data with reconciliation status

        Returns:
            BytesIO buffer with Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Conciliação"

        # Title
        ws.merge_cells("A1:H1")
        ws["A1"] = f"Relatório de Conciliação - Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Headers
        headers = [
            "ID Transação",
            "ID Pix",
            "Data",
            "Cliente",
            "Valor Esperado",
            "Valor Recebido",
            "Status",
            "Observações",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self._header_font
            cell.fill = self._header_fill
            cell.border = self._border
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        row = 4
        for txn in transactions:
            ws.cell(row=row, column=1, value=str(txn["transaction_id"]))
            ws.cell(row=row, column=2, value=txn.get("pix_transaction_id", ""))
            ws.cell(row=row, column=3, value=txn["date"]).number_format = self._datetime_format
            ws.cell(row=row, column=4, value=txn["client_name"])

            expected = Decimal(txn["expected_amount"]) / Decimal(100)
            ws.cell(row=row, column=5, value=float(expected)).number_format = self._money_format

            received = txn.get("received_amount")
            if received:
                received_decimal = Decimal(received) / Decimal(100)
                ws.cell(row=row, column=6, value=float(received_decimal)).number_format = self._money_format
            else:
                ws.cell(row=row, column=6, value="Pendente")

            ws.cell(row=row, column=7, value=txn["reconciliation_status"])
            ws.cell(row=row, column=8, value=txn.get("notes", ""))

            # Apply borders and conditional formatting
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                cell.border = self._border

                # Highlight issues
                if txn["reconciliation_status"] == "divergente":
                    cell.fill = PatternFill(
                        start_color="fef3cd", end_color="fef3cd", fill_type="solid"
                    )
                elif txn["reconciliation_status"] == "pendente":
                    cell.fill = PatternFill(
                        start_color="f8d7da", end_color="f8d7da", fill_type="solid"
                    )

            row += 1

        # Auto-adjust column widths
        column_widths = [36, 25, 18, 25, 15, 15, 15, 30]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
